"""Excel 财务报表解析：输出 {项目名: {年份: 值}}"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

YEAR_RE = re.compile(r"(19|20)\d{2}")

STATEMENT_KEYWORDS = [
    ("资产负债表", ["资产负债", "负债表", "balance sheet", "financial position", "balance"]),
    ("利润表", ["利润", "损益", "income", "profit", "comprehensive", "earnings"]),
    ("现金流量表", ["现金流量", "cash flow", "cash flows", "statement of cash"]),
]


@dataclass
class ExcelStatement:
    statement_type: str                        # 资产负债表/利润表/现金流量表/未知
    source_file: str
    sheet_name: str
    data: dict[str, dict[str, str]] = field(default_factory=dict)
    # data: {项目名: {年份字符串: 原始值字符串}}

    @property
    def items(self) -> list[str]:
        return list(self.data.keys())


def detect_statement_type(*texts: str) -> str:
    """根据文本关键词判断报表类型"""
    joined = " ".join(t.lower() for t in texts if t)
    for st, keywords in STATEMENT_KEYWORDS:
        if any(k in joined for k in keywords):
            return st
    return "未知"


def extract_year(text: str) -> str | None:
    """从文本中提取 4 位年份，如 '2024年度' '2024/12/31' -> '2024'"""
    if text is None:
        return None
    m = YEAR_RE.search(str(text))
    return m.group(0) if m else None


# 元数据行白名单：这些标签是报表的辅助信息，不应当作财务项目
_METADATA_LABELS = {
    "报告期", "报表类型", "单位", "币种", "编制单位", "编制日期",
    "项目", "会计机构",
}


def _is_metadata_label(name: str) -> bool:
    """判断是否为元数据标签（报告期/报表类型/单位/会计…），不应计入项目"""
    name = str(name).strip()
    if not name:
        return False
    if name in _METADATA_LABELS:
        return True
    if any(k in name for k in ("会计", "编制", "负责人", "法定代表人", "主管")):
        return True
    return False


def _looks_numeric(text: str) -> bool:
    """判断文本是否像数值（允许逗号/百分号/负号），用于过滤非数值文本"""
    t = str(text).strip()
    if not t:
        return False
    t2 = t.replace(",", "").replace("，", "").replace("%", "")
    try:
        float(t2)
        return True
    except ValueError:
        return False


def _cell_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def parse_excel(path: str | Path) -> list[ExcelStatement]:
    """解析一个 Excel 文件，每个含有效报表结构的 sheet 输出一个 ExcelStatement"""
    path = Path(path)
    wb = load_workbook(path, data_only=True, read_only=True)
    statements: list[ExcelStatement] = []

    for ws in wb.worksheets:
        rows = [[_cell_str(c) for c in row] for row in ws.iter_rows(values_only=True)]
        if not rows:
            continue

        # 1) 找表头行：含 >=1 个年份、且首列多为文本的行（取前 10 行内年份数最多者）
        header_idx, year_cols = None, {}
        for i, row in enumerate(rows[:10]):
            cols = {}
            for j, cell in enumerate(row[1:], start=1):
                y = extract_year(cell)
                if y:
                    cols[j] = y
            if len(cols) > len(year_cols):
                header_idx, year_cols = i, cols
        if header_idx is None or not year_cols:
            continue  # 该 sheet 无年份表头，跳过

        # 2) 报表类型：sheet 名 + 表头以上文本
        head_texts = [ws.title] + [c for r in rows[: header_idx + 1] for c in r]
        st_type = detect_statement_type(*head_texts)

        # 3) 数据行：表头之下，首列为项目名，年份列取值
        #    仅采纳数值型单元格，跳过「一季报/母公司报表」等非数值文本，避免污染匹配
        data: dict[str, dict[str, str]] = {}
        for row in rows[header_idx + 1:]:
            item = row[0].strip() if row and row[0] else ""
            if not item or _is_metadata_label(item):
                continue
            values = {}
            for j, year in year_cols.items():
                if j < len(row) and row[j] != "" and _looks_numeric(row[j]):
                    values[year] = row[j]
            if values:
                data[item] = values

        if data:
            statements.append(
                ExcelStatement(
                    statement_type=st_type,
                    source_file=path.name,
                    sheet_name=ws.title,
                    data=data,
                )
            )

    wb.close()
    return statements
